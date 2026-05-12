from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .io import ensure_dir, write_csv, write_json


MEASUREMENT_SHEETS = {
    "Aggression time",
    "Average movement acceleration",
    "Average movement ang. velocity",
    "Average overall speed",
    "Average speed when awake",
    "Cage center occupancy",
    "Cage front half occupancy",
    "Cage in rack",
    "Climbing events",
    "Max movement speed",
    "Number of agression events",
    "Rearing events",
    "Sleep bouts",
    "Social Distance",
    "Time spent climbing",
    "Time spent drinking water",
    "Time spent eating",
    "Time spent in cage corners",
    "Time spent inactive",
    "Time spent rearing",
    "Total distance traveled",
    "Tube interaction time",
}

INVIVO_FIELDNAMES = [
    "study_id",
    "source_file",
    "sheet",
    "measurement",
    "group",
    "cage",
    "animal_id",
    "time_bin",
    "value",
]

STUDY_FIELDNAMES = [
    "study_id",
    "source_file",
    "file_name",
    "sheet_count",
    "measurement_sheet_count",
    "groups",
    "measurements",
    "animal_columns",
    "time_rows",
]

RAW_JSON_FIELDNAMES = [
    "study_name",
    "source_file",
    "device_uid",
    "mouse_tag",
    "group",
    "datetime_from",
    "datetime_to",
    "measurement",
    "value",
]

CAGE_JSON_FIELDNAMES = [
    "study_name",
    "source_file",
    "measurement",
    "unit",
    "group",
    "time_bin",
    "value",
]

ANALYSIS_BAR_FIELDNAMES = [
    "study_name",
    "source_file",
    "metric",
    "chart_name",
    "unit",
    "window_hours",
    "dose_time",
    "end_time",
    "group",
    "mean_value",
    "error_min",
    "error_max",
    "n",
]

ANALYSIS_SCATTER_FIELDNAMES = [
    "study_name",
    "source_file",
    "metric",
    "window_hours",
    "dose_time",
    "end_time",
    "group",
    "value",
]

ANALYSIS_SIG_FIELDNAMES = [
    "study_name",
    "source_file",
    "metric",
    "window_hours",
    "dose_time",
    "end_time",
    "group",
    "t_stat",
    "p_value",
    "significant",
    "direction",
]


def _study_id(path: Path) -> str:
    stem = path.stem
    match = re.search(r"_(\d{3,5})_study$", stem)
    if match:
        return f"study{match.group(1)}"
    digest = hashlib.sha1(str(path).encode("utf-8")).hexdigest()[:8]
    return f"study-{digest}"


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _numeric(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def extract_mouse_data(mouse_dir: Path, out_dir: Path) -> dict[str, Any]:
    if not mouse_dir.exists():
        raise FileNotFoundError(mouse_dir)
    ensure_dir(out_dir)

    all_rows: list[dict[str, Any]] = []
    study_rows: list[dict[str, Any]] = []

    for path in sorted(mouse_dir.rglob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        study_id = _study_id(path)
        groups_seen: set[str] = set()
        measurements_seen: list[str] = []
        animal_columns = 0
        time_rows = 0

        for ws in wb.worksheets:
            if ws.title not in MEASUREMENT_SHEETS:
                continue
            measurements_seen.append(ws.title)

            group_row = [_clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            cage_row = [_clean(c.value) for c in next(ws.iter_rows(min_row=2, max_row=2))]
            animal_row = [_clean(c.value) for c in next(ws.iter_rows(min_row=3, max_row=3))]
            cols = []
            for col_idx in range(2, ws.max_column + 1):
                group = group_row[col_idx - 1] if col_idx - 1 < len(group_row) else ""
                cage = cage_row[col_idx - 1] if col_idx - 1 < len(cage_row) else ""
                animal = animal_row[col_idx - 1] if col_idx - 1 < len(animal_row) else ""
                if group or cage or animal:
                    cols.append((col_idx, group, cage, animal))
                    if group:
                        groups_seen.add(group)

            animal_columns = max(animal_columns, len(cols))
            for row in ws.iter_rows(min_row=5, values_only=True):
                time_bin = _clean(row[0] if row else "")
                if not time_bin:
                    continue
                time_rows += 1
                for col_idx, group, cage, animal in cols:
                    value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                    numeric_value = _numeric(value)
                    if numeric_value is None:
                        continue
                    all_rows.append(
                        {
                            "study_id": study_id,
                            "source_file": str(path),
                            "sheet": ws.title,
                            "measurement": ws.title,
                            "group": group,
                            "cage": cage,
                            "animal_id": animal,
                            "time_bin": time_bin,
                            "value": numeric_value,
                        }
                    )

        if measurements_seen:
            study_rows.append(
                {
                    "study_id": study_id,
                    "source_file": str(path),
                    "file_name": path.name,
                    "sheet_count": len(wb.sheetnames),
                    "measurement_sheet_count": len(measurements_seen),
                    "groups": ";".join(sorted(groups_seen)),
                    "measurements": ";".join(measurements_seen),
                    "animal_columns": animal_columns,
                    "time_rows": time_rows,
                }
            )

    json_summary = _extract_olden_json(mouse_dir / "olden-autopulled", out_dir)

    write_csv(out_dir / "invivo_measurements_long.csv", all_rows, INVIVO_FIELDNAMES)
    write_csv(out_dir / "invivo_studies_inventory.csv", study_rows, STUDY_FIELDNAMES)
    summary = {
        "source_dir": str(mouse_dir),
        "excel_studies": len(study_rows),
        "excel_measurement_rows": len(all_rows),
        **json_summary,
        "outputs": {
            "excel_measurements": str(out_dir / "invivo_measurements_long.csv"),
            "excel_inventory": str(out_dir / "invivo_studies_inventory.csv"),
            "raw_json_measurements": str(out_dir / "invivo_olden_raw_measurements_long.csv"),
            "cage_in_rack_json": str(out_dir / "invivo_olden_cage_in_rack.csv"),
            "analysis_bars": str(out_dir / "invivo_olden_analysis_bars.csv"),
            "analysis_scatter": str(out_dir / "invivo_olden_analysis_scatter.csv"),
            "analysis_significance": str(out_dir / "invivo_olden_analysis_significance.csv"),
        },
    }
    write_json(out_dir / "invivo_extract_summary.json", summary)
    return summary


def _extract_olden_json(olden_dir: Path, out_dir: Path) -> dict[str, Any]:
    raw_rows: list[dict[str, Any]] = []
    cage_rows: list[dict[str, Any]] = []
    bar_rows: list[dict[str, Any]] = []
    scatter_rows: list[dict[str, Any]] = []
    sig_rows: list[dict[str, Any]] = []
    json_studies: set[str] = set()

    if not olden_dir.exists():
        return {
            "olden_json_studies": 0,
            "olden_raw_measurement_rows": 0,
            "olden_cage_in_rack_rows": 0,
            "olden_analysis_bar_rows": 0,
            "olden_analysis_scatter_rows": 0,
            "olden_analysis_significance_rows": 0,
        }

    for path in sorted(olden_dir.rglob("*.json")):
        study_name = path.parent.name
        json_studies.add(study_name)
        try:
            obj = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue

        if path.name == "raw_data.json":
            group_by_device = _metadata_group_by_device(path.parent / "metadata.json")
            for device_uid, records in obj.items() if isinstance(obj, dict) else []:
                if not isinstance(records, list):
                    continue
                for record in records:
                    if not isinstance(record, dict):
                        continue
                    for key, value in record.items():
                        if key in {
                            "datetime_from",
                            "datetime_to",
                            "device_uid",
                            "mouse_tag",
                            "empty_food",
                            "agression",
                            "dfi_age",
                        }:
                            continue
                        if isinstance(value, bool):
                            continue
                        numeric_value = _numeric(value)
                        if numeric_value is None:
                            continue
                        raw_rows.append(
                            {
                                "study_name": study_name,
                                "source_file": str(path),
                                "device_uid": record.get("device_uid", device_uid),
                                "mouse_tag": record.get("mouse_tag", ""),
                                "group": group_by_device.get(device_uid, ""),
                                "datetime_from": record.get("datetime_from", ""),
                                "datetime_to": record.get("datetime_to", ""),
                                "measurement": key,
                                "value": numeric_value,
                            }
                        )

        elif path.name == "cage_in_rack.json" and isinstance(obj, dict):
            measurement = obj.get("name", "cage_in_rack")
            unit = obj.get("unit", "")
            timepoints = obj.get("timepoints", [])
            for group_obj in obj.get("groups", []):
                group = group_obj.get("label", "")
                for time_bin, value in zip(timepoints, group_obj.get("data", [])):
                    numeric_value = _numeric(value)
                    if numeric_value is None:
                        continue
                    cage_rows.append(
                        {
                            "study_name": study_name,
                            "source_file": str(path),
                            "measurement": measurement,
                            "unit": unit,
                            "group": group,
                            "time_bin": time_bin,
                            "value": numeric_value,
                        }
                    )

        elif path.name == "analysis.json" and isinstance(obj, dict):
            metric = obj.get("metric", "")
            chart_name = obj.get("chart_name", "")
            unit = obj.get("unit", "")
            for window in obj.get("windows", []):
                window_hours = window.get("window_hours", "")
                dose_time = window.get("dose_time", "")
                end_time = window.get("end_time", "")
                bar_data = window.get("bar_data", {})
                for bar in bar_data.get("bars", []):
                    bar_rows.append(
                        {
                            "study_name": study_name,
                            "source_file": str(path),
                            "metric": metric or bar_data.get("metric", ""),
                            "chart_name": chart_name or bar_data.get("chart_name", ""),
                            "unit": unit or bar_data.get("unit", ""),
                            "window_hours": window_hours,
                            "dose_time": dose_time,
                            "end_time": end_time,
                            "group": bar.get("label", ""),
                            "mean_value": bar.get("value", ""),
                            "error_min": bar.get("error_min", ""),
                            "error_max": bar.get("error_max", ""),
                            "n": bar.get("n", ""),
                        }
                    )
                for point in bar_data.get("scatter", []):
                    scatter_rows.append(
                        {
                            "study_name": study_name,
                            "source_file": str(path),
                            "metric": metric or bar_data.get("metric", ""),
                            "window_hours": window_hours,
                            "dose_time": dose_time,
                            "end_time": end_time,
                            "group": point.get("group", ""),
                            "value": point.get("y", ""),
                        }
                    )
                for group, stats in window.get("significance", {}).items():
                    sig_rows.append(
                        {
                            "study_name": study_name,
                            "source_file": str(path),
                            "metric": metric,
                            "window_hours": window_hours,
                            "dose_time": dose_time,
                            "end_time": end_time,
                            "group": group,
                            "t_stat": stats.get("t_stat", ""),
                            "p_value": stats.get("p_value", ""),
                            "significant": stats.get("significant", ""),
                            "direction": stats.get("direction", ""),
                        }
                    )

    write_csv(out_dir / "invivo_olden_raw_measurements_long.csv", raw_rows, RAW_JSON_FIELDNAMES)
    write_csv(out_dir / "invivo_olden_cage_in_rack.csv", cage_rows, CAGE_JSON_FIELDNAMES)
    write_csv(out_dir / "invivo_olden_analysis_bars.csv", bar_rows, ANALYSIS_BAR_FIELDNAMES)
    write_csv(out_dir / "invivo_olden_analysis_scatter.csv", scatter_rows, ANALYSIS_SCATTER_FIELDNAMES)
    write_csv(out_dir / "invivo_olden_analysis_significance.csv", sig_rows, ANALYSIS_SIG_FIELDNAMES)
    return {
        "olden_json_studies": len(json_studies),
        "olden_raw_measurement_rows": len(raw_rows),
        "olden_cage_in_rack_rows": len(cage_rows),
        "olden_analysis_bar_rows": len(bar_rows),
        "olden_analysis_scatter_rows": len(scatter_rows),
        "olden_analysis_significance_rows": len(sig_rows),
    }


def _metadata_group_by_device(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    try:
        metadata = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    candidates: dict[str, set[str]] = {}
    for group in metadata.get("groups", []):
        group_name = group.get("name", "")
        for cage in group.get("cages", []):
            for uid in cage.get("device_uids", []):
                candidates.setdefault(uid, set()).add(group_name)
    return {uid: next(iter(names)) for uid, names in candidates.items() if len(names) == 1}
