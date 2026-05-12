from __future__ import annotations

import csv
import hashlib
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .io import ensure_dir, write_csv, write_json


ASSAY_FIELDNAMES = [
    "assay_record_id",
    "peptide_id",
    "source_sheet",
    "location",
    "date",
    "status",
    "producer",
    "cell_line",
    "receptor",
    "modification",
    "compound",
    "assay",
    "position",
    "ec50_nm",
    "fold_vs_ref",
    "pec50",
    "bias_temporary",
    "bias",
    "emax_pct",
    "hill_slope",
    "bottom",
    "pkb",
    "notes",
    "condition_group_id",
]

PEPTIDE_FIELDNAMES = [
    "peptide_id",
    "compound",
    "modification",
    "receptor_family",
    "first_seen_date",
    "assay_count",
    "receptors",
    "assays",
    "producers",
]

QC_FIELDNAMES = [
    "status",
    "location",
    "date",
    "receptor",
    "assay",
    "raw_file",
    "rows_total",
    "rows_promoted",
    "reference",
    "ref_ec50_nm",
    "reason",
]


def _slug(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "-", value.strip()).strip("-")
    return value.lower() or "unknown"


def _clean_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace(" ", "_").replace("-", "_")
    text = re.sub(r"[^a-z0-9_]+", "", text)
    return text


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _peptide_id(compound: str, modification: str) -> str:
    key = f"{compound}|{modification}".encode("utf-8")
    return "PEP-" + hashlib.sha1(key).hexdigest()[:10].upper()


def _record_id(sheet: str, row_number: int, compound: str, receptor: str, assay: str) -> str:
    key = f"{sheet}|{row_number}|{compound}|{receptor}|{assay}".encode("utf-8")
    return "ASSAY-" + hashlib.sha1(key).hexdigest()[:12].upper()


def _family_from_sheet(sheet_name: str) -> str:
    upper = sheet_name.upper()
    if "NPS" in upper:
        return "NPS"
    if "OXN" in upper or "OREXIN" in upper:
        return "OXN"
    if "MCH" in upper:
        return "MCH"
    return _slug(sheet_name).upper()


def _condition_group(receptor_family: str, cell_line: str, receptor: str, assay: str) -> str:
    parts = [_slug(receptor_family), _slug(cell_line), _slug(receptor), _slug(assay)]
    return "CG-" + "-".join(parts)


def ingest_mastersheet(xlsx_path: Path, out_dir: Path) -> dict[str, Any]:
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    ensure_dir(out_dir)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    assay_rows: list[dict[str, Any]] = []
    qc_rows: list[dict[str, Any]] = []
    peptide_seen: dict[str, dict[str, Any]] = {}
    peptide_sets: dict[str, dict[str, set[str]]] = defaultdict(
        lambda: {"receptors": set(), "assays": set(), "producers": set()}
    )

    for ws in wb.worksheets:
        headers = [_clean_header(cell.value) for cell in next(ws.iter_rows(max_row=1))]
        if ws.title == "Internal Plate QC":
            for row in ws.iter_rows(min_row=2, values_only=True):
                values = dict(zip(headers, row))
                qc_rows.append(
                    {
                        "status": _clean_cell(values.get("status")),
                        "location": _clean_cell(values.get("location")),
                        "date": _clean_cell(values.get("date")),
                        "receptor": _clean_cell(values.get("receptor")),
                        "assay": _clean_cell(values.get("assay")),
                        "raw_file": _clean_cell(values.get("raw_file")),
                        "rows_total": _clean_cell(values.get("rows_total")),
                        "rows_promoted": _clean_cell(values.get("rows_promoted")),
                        "reference": _clean_cell(values.get("reference")),
                        "ref_ec50_nm": _clean_cell(values.get("ref_ec50_nm")),
                        "reason": _clean_cell(values.get("reason")),
                    }
                )
            continue

        receptor_family = _family_from_sheet(ws.title)
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = dict(zip(headers, row))
            compound = _clean_cell(values.get("compound"))
            modification = _clean_cell(values.get("modification"))
            receptor = _clean_cell(values.get("receptor"))
            assay = _clean_cell(values.get("assay"))
            if not compound and not modification and not receptor and not assay:
                continue

            peptide_id = _peptide_id(compound, modification)
            date = _clean_cell(values.get("date"))
            record = {
                "assay_record_id": _record_id(ws.title, row_number, compound, receptor, assay),
                "peptide_id": peptide_id,
                "source_sheet": ws.title,
                "location": _clean_cell(values.get("location")),
                "date": date,
                "status": _clean_cell(values.get("status")),
                "producer": _clean_cell(values.get("producer")),
                "cell_line": _clean_cell(values.get("cell_line")),
                "receptor": receptor,
                "modification": modification,
                "compound": compound,
                "assay": assay,
                "position": _clean_cell(values.get("position")),
                "ec50_nm": _clean_cell(values.get("ec50_nm")),
                "fold_vs_ref": _clean_cell(values.get("fold_vs_ref")),
                "pec50": _clean_cell(values.get("pec50")),
                "bias_temporary": _clean_cell(values.get("bias_temporary")),
                "bias": _clean_cell(values.get("bias")),
                "emax_pct": _clean_cell(values.get("emax_pct")),
                "hill_slope": _clean_cell(values.get("nh")),
                "bottom": _clean_cell(values.get("bottom")),
                "pkb": _clean_cell(values.get("pkb")),
                "notes": _clean_cell(values.get("notes")),
                "condition_group_id": _condition_group(
                    receptor_family,
                    _clean_cell(values.get("cell_line")),
                    receptor,
                    assay,
                ),
            }
            assay_rows.append(record)

            if peptide_id not in peptide_seen:
                peptide_seen[peptide_id] = {
                    "peptide_id": peptide_id,
                    "compound": compound,
                    "modification": modification,
                    "receptor_family": receptor_family,
                    "first_seen_date": date,
                    "assay_count": 0,
                }
            peptide_seen[peptide_id]["assay_count"] += 1
            if date and (
                not peptide_seen[peptide_id]["first_seen_date"]
                or date < peptide_seen[peptide_id]["first_seen_date"]
            ):
                peptide_seen[peptide_id]["first_seen_date"] = date
            peptide_sets[peptide_id]["receptors"].add(receptor)
            peptide_sets[peptide_id]["assays"].add(assay)
            producer = _clean_cell(values.get("producer"))
            if producer:
                peptide_sets[peptide_id]["producers"].add(producer)

    peptide_rows = []
    for peptide_id, row in sorted(peptide_seen.items()):
        sets = peptide_sets[peptide_id]
        peptide_rows.append(
            {
                **row,
                "receptors": ";".join(sorted(x for x in sets["receptors"] if x)),
                "assays": ";".join(sorted(x for x in sets["assays"] if x)),
                "producers": ";".join(sorted(x for x in sets["producers"] if x)),
            }
        )

    write_csv(out_dir / "invitro_assays.csv", assay_rows, ASSAY_FIELDNAMES)
    write_csv(out_dir / "peptides.csv", peptide_rows, PEPTIDE_FIELDNAMES)
    write_csv(out_dir / "plate_qc.csv", qc_rows, QC_FIELDNAMES)

    summary = {
        "source": str(xlsx_path),
        "sheets": wb.sheetnames,
        "assay_records": len(assay_rows),
        "peptides": len(peptide_rows),
        "qc_records": len(qc_rows),
        "outputs": {
            "invitro_assays": str(out_dir / "invitro_assays.csv"),
            "peptides": str(out_dir / "peptides.csv"),
            "plate_qc": str(out_dir / "plate_qc.csv"),
        },
    }
    write_json(out_dir / "ingest_summary.json", summary)
    return summary


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))

